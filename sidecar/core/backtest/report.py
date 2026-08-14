"""
Backtest output: a spreadsheet and a credibility argument (§11.2 Phase 1).

Three writers over the same BacktestReport:

  write_workbook  a multi-sheet .xlsx, when openpyxl is available
  write_csvs      the same tables as CSV, so the harness has no hard dependency
  render_markdown the credibility argument, caveats included

The caveats are not decoration. This run validates the *mechanism* — GTS read ->
mapper -> evidence extraction -> band engine -> comparison — on a population that
§6a establishes is far too thin to say anything about adjudication quality. A
report that leads with an agreement rate and buries that would be misleading, so
the caveats are generated from the run's own numbers and sit alongside the
headline, not in an appendix.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Optional

from .harness import BacktestReport, CaseOutcome, Outcome

CASE_COLUMNS: list[tuple[str, str]] = [
    ("case_id", "Case ID"),
    ("bp_id", "BP"),
    ("bp_name", "BP name"),
    ("bp_country", "BP country"),
    ("legal_regulation", "Legal regulation"),
    ("address_id", "Address ID"),
    ("spl_entry_id", "SPL entry"),
    ("spl_entry_name", "SPL entry name"),
    ("spl_entry_address", "SPL entry address"),
    ("spl_list_type", "List type"),
    ("spl_programme", "Programme"),
    ("match_basis", "Match basis"),
    ("release_reason_text", "Human release reason"),
    ("decided_by", "Decided by"),
]


def _case_row(o: CaseOutcome) -> dict[str, object]:
    row: dict[str, object] = {label: getattr(o, attr) for attr, label in CASE_COLUMNS}
    row["Human verdict"] = o.human_verdict.value
    row["Expected band"] = o.expected_label
    row["Actual band"] = o.actual_band.value if o.actual_band else ""
    row["Outcome"] = o.outcome.value
    row["Evidence items"] = len(o.ledger.items)
    row["Evidence available"] = o.evidence_available
    counts = o.category_counts
    for category in (
        "DISP_EXCL", "STRONG_DISC", "WEAK_DISC", "NEUTRAL",
        "WEAK_CORR", "STRONG_CORR", "DISP_CONF",
    ):
        row[category] = counts.get(category, 0)
    row["Elapsed ms"] = o.elapsed_ms
    row["Error"] = o.error
    row["Rationale"] = o.rationale
    return row


def _evidence_rows(report: BacktestReport) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for o in report.outcomes:
        for idx, item in enumerate(o.ledger.items, 1):
            rows.append({
                "Case ID": o.case_id,
                "BP": o.bp_id,
                "Sort order": idx,
                "Category": item.category.value,
                "Data element": item.data_element,
                "BP value": item.bp_value,
                "SPL value": item.spl_value,
                "Data available": item.data_available,
                "Assessment": item.assessment,
            })
    return rows


def _skipped_rows(report: BacktestReport) -> list[dict[str, object]]:
    return [{
        "Record key": s.record_key,
        "BP": s.bp_id,
        "BP name": s.bp_name,
        "Legal regulation": s.legal_regulation,
        "Address ID": s.address_id,
        "Human release reason": s.release_reason_text,
        "Human verdict": s.human_verdict.value,
        "Why not scored": s.skip_reason.value,
    } for s in report.skipped]


def _summary_rows(report: BacktestReport) -> list[tuple[str, object]]:
    rate = report.agreement_rate
    non_contra = report.non_contradiction_rate
    return [
        ("Generated", report.generated_at.isoformat(timespec="seconds")),
        ("Source", report.source),
        ("Live read from AGP", "yes" if report.live_read else "no (replayed from file)"),
        ("Extractor", "MOCK — not the LLM" if report.mock_extractor else "LLM"),
        ("Legal regulations", ", ".join(report.regulations)),
        ("Model", report.model_version),
        ("Prompt version", report.prompt_version),
        ("Band logic version", report.band_logic_version),
        ("Taxonomy version", report.taxonomy_version),
        ("", ""),
        ("SAFETY GATE (§8)", "PASS" if report.safety_gate_passed else "FAIL"),
        ("Confirmed matches proposed for clearing", len(report.safety_violations)),
        ("Confirmed-match pairs adjudicated", report.true_positive_pairs),
        ("", ""),
        ("Decided records read", report.records_read),
        ("Records scored", report.records_scored),
        ("Records skipped", len(report.skipped)),
        ("Pairs adjudicated", report.pairs_adjudicated),
        ("", ""),
        ("Agreement rate", "n/a" if rate is None else f"{rate:.1%}"),
        ("Non-contradiction rate", "n/a" if non_contra is None else f"{non_contra:.1%}"),
        *[(f"Outcome — {k}", v) for k, v in report.outcome_counts.items()],
        ("", ""),
        *[(f"Band — {k}", v) for k, v in report.band_counts.items()],
        ("", ""),
        *[(f"Skipped — {k}", v) for k, v in report.skip_counts.items()],
    ]


# ------------------------------------------------------------------- caveats

#: Established facts about the AGP population, from the integration notes. These
#: bound what any run against AGP can demonstrate, regardless of its numbers.
STATIC_CAVEATS: list[str] = [
    "No BP discriminators exist anywhere in AGP: of all 4,119 business partners, "
    "**0** carry a date of birth, nationality, birthplace, name at birth, "
    "foundation date or industry (§6a, surveyed in full, not sampled).",
    "`BusinessPartnerCategory` reads `2` (Organization) for every record sampled, "
    "including partners plainly named after people, so entity type — normally the "
    "most common dispositive exclusion (§3.2) — is not a usable discriminator on "
    "this data.",
    "The SPL entry's **name and address** are available and are used (§3a, carried "
    "as HTML in `MatchedName`/`MatchedAddress`). Aliases, DOB, place of birth, "
    "nationality, passport/national ID/registration/LEI, listing history and "
    "remarks are **not** — they need a Z CDS view over `/SAPSLL/TSPL*`, which "
    "needs client 200 (§4).",
    "GTS exposes no similarity percentage on this service, so `match_percentage` "
    "is reported as 0.0 meaning \"not reported\" and must never be rendered as a "
    "confidence (§3, §6.4).",
    "**Auto-clear is structurally unreachable in this harness.** It requires a "
    "matching precedent (§3.4) and the precedent store is an in-memory stub, so "
    "the best band a correctly-cleared case can reach is Propose Clear. A zero in "
    "the Auto-clear column is an artefact of that, not a finding about the data.",
    "Network evidence (§3.1 #8) is unavailable: identifications, associated "
    "blocked objects and associated banks all come back empty, and GTS's own "
    "`*IsHidden` flags confirm those sections are empty rather than unread (§6a).",
]


def dynamic_caveats(report: BacktestReport) -> list[str]:
    """Caveats computed from this run, so the numbers can't drift from the prose."""
    out: list[str] = []

    if report.mock_extractor:
        out.append(
            "**This run used the mock extractor, not the LLM.** The mock emits a "
            "fixed entity-type/country ledger, so the bands below exercise the read "
            "path, the mapper and the band engine only — they say nothing whatsoever "
            "about evidence quality. Rerun without `--mock` before quoting any "
            "number here."
        )

    if not report.live_read:
        out.append(
            "Records were replayed from a saved file rather than read live from AGP, "
            "so this run proves the reasoning path but not current connectivity."
        )

    if report.records_read == 0:
        out.append(
            "**No decided records were read at all.** Either the filter matched "
            "nothing or the read failed — this run demonstrates nothing."
        )
        return out

    no_detail = report.skip_counts.get(
        "No SPL hit detail on the record — nothing to adjudicate", 0
    )
    if no_detail:
        out.append(
            f"{no_detail} of {report.records_read} decided records carry no SPL hit "
            "detail and could not be run. A block on a licence, embargo or customs "
            "regulation has no sanctioned party to compare against, so there is no "
            "identity question — this is GTS being correct, not data loss (§6a)."
        )

    if report.pairs_adjudicated == 0:
        out.append(
            "**Nothing was adjudicated**, so there is no agreement rate and no "
            "credibility argument to make. The harness ran; the population was empty."
        )
        return out

    if report.pairs_adjudicated < 10:
        out.append(
            f"Only {report.pairs_adjudicated} BP↔entry pair(s) were adjudicated. "
            "At this size a single case moves the agreement rate by tens of "
            "percentage points, so **no rate below should be read as a "
            "performance measure.** This run validates the mechanism end to end; "
            "quality needs either the Z view (client 200) or a real client extract."
        )

    abstentions = report.outcome_counts[Outcome.ABSTENTION.value]
    if abstentions and abstentions == report.scorable_pairs:
        out.append(
            f"Every scorable pair ({abstentions}) came back **Review**. That is the "
            "band engine behaving as specified rather than failing: with no SPL "
            "entry content and no BP discriminators, the ledger is thin, missing "
            "data is neutral (§3.1 #3), and neutral can never clear a hit. It also "
            "means the run says nothing about whether the reasoning is *good*."
        )
    elif abstentions:
        out.append(
            f"{abstentions} pair(s) came back Review — the agent declined to take a "
            "position. Counted separately from both agreement and divergence, "
            "because on this data Review is the designed outcome for a thin file."
        )

    if report.errors:
        out.append(
            f"{len(report.errors)} pair(s) failed during evidence extraction and are "
            "counted as errors, never as agreements. See the Error column."
        )

    if not report.true_positive_pairs:
        out.append(
            "**No confirmed-match records were adjudicated, so the safety gate was "
            "never exercised.** A PASS here means \"not violated\", not \"tested\": "
            "the metric that matters (§8) needs records the reviewer labelled "
            "\"Business partner matched to SPL record\" *and* that carry hit detail."
        )

    return out


# ------------------------------------------------------------------- markdown


def _matrix_markdown(report: BacktestReport) -> list[str]:
    rows, cols, cells = report.confusion_matrix()
    lines = ["| Human verdict → agent band | " + " | ".join(cols) + " |",
             "|---|" + "|".join("---" for _ in cols) + "|"]
    for r in rows:
        lines.append(f"| {r} | " + " | ".join(str(cells[(r, c)]) for c in cols) + " |")
    return lines


def _ledger_markdown(o: CaseOutcome) -> list[str]:
    if not o.ledger.items:
        return ["", "_Empty evidence ledger._"]
    lines = ["", "| Category | Data element | BP value | SPL value | Available | Assessment |",
             "|---|---|---|---|---|---|"]
    for item in o.ledger.items:
        lines.append(
            f"| {item.category.value} | {item.data_element} | "
            f"{item.bp_value or '—'} | {item.spl_value or '—'} | "
            f"{'yes' if item.data_available else 'NO'} | {item.assessment} |"
        )
    return lines


def render_markdown(report: BacktestReport) -> str:
    rate = report.agreement_rate
    non_contra = report.non_contradiction_rate
    L: list[str] = []

    L.append("# Phase 1 backtest — SPL hit adjudication")
    L.append("")
    L.append(f"Run {report.generated_at.isoformat(timespec='seconds')} against "
             f"**{report.source or 'unspecified source'}**.")
    L.append("")
    L.append(f"Regulations `{', '.join(report.regulations) or 'all'}` · "
             f"model `{report.model_version}` · prompt `{report.prompt_version}` · "
             f"band logic `{report.band_logic_version}` · "
             f"taxonomy `{report.taxonomy_version}`")
    L.append("")
    L.append("Ground truth is the release reason a human reviewer picked in GTS. "
             "Read-only: GETs against `LLS_BPADDR_MNG_SRV`, no writeback, and the "
             "release entities §5.4 reserves for a human were never addressed.")
    L.append("")

    # The gate first — it is the only thing that can fail the run.
    L.append("## The metric that matters (§8)")
    L.append("")
    if report.safety_gate_passed:
        L.append(f"**PASS** — zero of {report.true_positive_pairs} confirmed-match "
                 "pair(s) came back Auto-clear or Propose Clear.")
        if not report.true_positive_pairs:
            L.append("")
            L.append("> Read this as *not violated*, not as *tested*. No record "
                     "labelled \"Business partner matched to SPL record\" reached "
                     "the pipeline, so the gate was never exercised.")
    else:
        L.append(f"**FAIL — {len(report.safety_violations)} confirmed match(es) came "
                 "back as a clear.** The band logic is broken and nothing else in "
                 "this report matters until it is fixed.")
        for o in report.safety_violations:
            L.append("")
            L.append(f"- `{o.case_id}` — BP {o.bp_id} {o.bp_name!r} vs entry "
                     f"{o.spl_entry_id} {o.spl_entry_name!r} → **{o.actual_band.value}** "
                     f"(human: {o.release_reason_text!r})")
    L.append("")

    L.append("## Population")
    L.append("")
    L.append("| | |")
    L.append("|---|---|")
    L.append(f"| Decided records read | {report.records_read} |")
    L.append(f"| Records scored | {report.records_scored} |")
    L.append(f"| Records skipped | {len(report.skipped)} |")
    L.append(f"| BP↔entry pairs adjudicated | {report.pairs_adjudicated} |")
    L.append("")
    if report.skip_counts:
        L.append("Skipped, by reason:")
        L.append("")
        for reason, n in sorted(report.skip_counts.items(), key=lambda kv: -kv[1]):
            L.append(f"- **{n}** — {reason}")
        L.append("")

    L.append("## Results")
    L.append("")
    L.append(f"- Agreement rate: **{'n/a' if rate is None else f'{rate:.1%}'}** "
             f"({report.outcome_counts[Outcome.AGREEMENT.value]} of "
             f"{report.scorable_pairs} scorable pairs)")
    L.append(f"- Non-contradiction rate (agreement + abstention): "
             f"**{'n/a' if non_contra is None else f'{non_contra:.1%}'}**")
    L.append("")
    for name, n in report.outcome_counts.items():
        L.append(f"- {name}: {n}")
    L.append("")
    L.append("Bands returned: " + (", ".join(
        f"{k} {v}" for k, v in report.band_counts.items()) or "none"))
    L.append("")

    L.append("## Confusion matrix")
    L.append("")
    L.extend(_matrix_markdown(report))
    L.append("")

    L.append("## Divergences")
    L.append("")
    if not report.divergences:
        L.append("None — no case contradicted its reviewer.")
        L.append("")
    else:
        L.append("§11.2 says to investigate each one. Some will be human error, "
                 "which is a finding worth having.")
        L.append("")
        for o in report.divergences:
            L.append(f"### `{o.case_id}` — {o.outcome.value}")
            L.append("")
            L.append(f"- BP {o.bp_id} — {o.bp_name!r} ({o.bp_country or 'country n/a'})")
            L.append(f"- SPL entry {o.spl_entry_id} — {o.spl_entry_name or '(no name)'!r}"
                     + (f", {o.spl_entry_address}" if o.spl_entry_address else ""))
            L.append(f"- List `{o.spl_list_type}` · programme `{o.spl_programme}` · "
                     f"severity `{o.list_severity.value}`")
            L.append(f"- Match basis: {o.match_basis}")
            L.append(f"- Human: {o.release_reason_text!r} → {o.human_verdict.value}"
                     + (f" (decided by {o.decided_by})" if o.decided_by else ""))
            L.append(f"- Expected **{o.expected_label}**, agent returned "
                     f"**{o.actual_band.value if o.actual_band else 'ERROR'}**")
            L.extend(_ledger_markdown(o))
            L.append("")

    if report.errors:
        L.append("## Errors")
        L.append("")
        for o in report.errors:
            L.append(f"- `{o.case_id}` — {o.error}")
        L.append("")

    L.append("## What this run does and does not show")
    L.append("")
    for c in dynamic_caveats(report):
        L.append(f"- {c}")
    for c in STATIC_CAVEATS:
        L.append(f"- {c}")
    L.append("")
    exercised = ["the mapper", "the band engine",
                 "the comparison against real human decisions"]
    exercised.insert(0, "the live GTS read path" if report.live_read
                     else "a replay of previously-read GTS records")
    exercised.insert(
        2,
        "the mock extractor (**not** the LLM evidence extractor)"
        if report.mock_extractor else "the LLM evidence extractor",
    )
    L.append("In short: **this validates the mechanism, not the adjudication.** "
             "What ran end to end here: " + ", ".join(exercised) + ". Judging whether "
             "the reasoning is *good* needs realistic data on both sides — the Z view "
             "over `/SAPSLL/TSPL*` (client 200), a real client extract, or seeded BPs "
             "in AGP (§6a).")
    L.append("")

    return "\n".join(L)


# --------------------------------------------------------------------- writers


def write_csvs(report: BacktestReport, out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []

    summary = out_dir / "summary.csv"
    with summary.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Metric", "Value"])
        w.writerows(_summary_rows(report))
    written.append(summary)

    cases = out_dir / "cases.csv"
    rows = [_case_row(o) for o in report.outcomes]
    with cases.open("w", newline="", encoding="utf-8") as f:
        if rows:
            w = csv.DictWriter(f, fieldnames=list(rows[0]))
            w.writeheader()
            w.writerows(rows)
        else:
            csv.writer(f).writerow(["(no cases adjudicated)"])
    written.append(cases)

    evidence = out_dir / "evidence.csv"
    ev_rows = _evidence_rows(report)
    with evidence.open("w", newline="", encoding="utf-8") as f:
        if ev_rows:
            w = csv.DictWriter(f, fieldnames=list(ev_rows[0]))
            w.writeheader()
            w.writerows(ev_rows)
        else:
            csv.writer(f).writerow(["(no evidence items)"])
    written.append(evidence)

    skipped = out_dir / "skipped.csv"
    sk_rows = _skipped_rows(report)
    with skipped.open("w", newline="", encoding="utf-8") as f:
        if sk_rows:
            w = csv.DictWriter(f, fieldnames=list(sk_rows[0]))
            w.writeheader()
            w.writerows(sk_rows)
        else:
            csv.writer(f).writerow(["(nothing skipped)"])
    written.append(skipped)

    return written


def write_workbook(report: BacktestReport, path: Path) -> Optional[Path]:
    """
    Write the multi-sheet .xlsx. Returns None if openpyxl isn't installed, so
    the CSVs remain the guaranteed output and the harness keeps no hard
    dependency on it.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="354A5F")
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()

    def add_sheet(title: str, columns: list[str], rows: list[list[object]],
                  widths: Optional[dict[int, int]] = None):
        ws = wb.create_sheet(title)
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        for idx in range(1, len(columns) + 1):
            width = (widths or {}).get(idx, 18)
            ws.column_dimensions[get_column_letter(idx)].width = width
        return ws

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for k, v in _summary_rows(report):
        ws.append([k, v])
    for row in ws.iter_rows(min_row=2, max_col=1):
        label = row[0].value or ""
        if label.startswith("SAFETY GATE"):
            row[0].font = Font(bold=True)
            ws.cell(row=row[0].row, column=2).font = Font(
                bold=True, color="1E7B34" if report.safety_gate_passed else "C00000"
            )
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 60

    # Confusion matrix on the summary sheet, below the metrics
    rows, cols, cells = report.confusion_matrix()
    ws.append([])
    ws.append(["Confusion matrix: human verdict (rows) x agent band (columns)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Human verdict"] + cols)
    for r in rows:
        ws.append([r] + [cells[(r, c)] for c in cols])

    case_rows = [_case_row(o) for o in report.outcomes]
    if case_rows:
        columns = list(case_rows[0])
        add_sheet(
            "Cases", columns, [[r[c] for c in columns] for r in case_rows],
            widths={columns.index("Rationale") + 1: 70,
                    columns.index("BP name") + 1: 28,
                    columns.index("SPL entry name") + 1: 28,
                    columns.index("Human release reason") + 1: 44,
                    columns.index("Match basis") + 1: 34},
        )
    else:
        add_sheet("Cases", ["(no cases adjudicated)"], [])

    ev_rows = _evidence_rows(report)
    if ev_rows:
        columns = list(ev_rows[0])
        add_sheet(
            "Evidence", columns, [[r[c] for c in columns] for r in ev_rows],
            widths={columns.index("Assessment") + 1: 80,
                    columns.index("Data element") + 1: 26},
        )
    else:
        add_sheet("Evidence", ["(no evidence items)"], [])

    div_columns = ["Case ID", "Outcome", "BP", "BP name", "SPL entry",
                   "SPL entry name", "Human release reason", "Expected band",
                   "Actual band", "Match basis", "Rationale"]
    add_sheet("Divergences", div_columns, [[
        o.case_id, o.outcome.value, o.bp_id, o.bp_name, o.spl_entry_id,
        o.spl_entry_name, o.release_reason_text, o.expected_label,
        o.actual_band.value if o.actual_band else "", o.match_basis, o.rationale,
    ] for o in report.divergences], widths={11: 70, 7: 44})

    sk_rows = _skipped_rows(report)
    if sk_rows:
        columns = list(sk_rows[0])
        add_sheet("Skipped", columns,
                  [[r[c] for c in columns] for r in sk_rows],
                  widths={columns.index("Why not scored") + 1: 60})
    else:
        add_sheet("Skipped", ["(nothing skipped)"], [])

    caveat_ws = add_sheet("Caveats", ["#", "Caveat"], [
        [i, c] for i, c in enumerate(dynamic_caveats(report) + STATIC_CAVEATS, 1)
    ], widths={1: 6, 2: 140})
    for row in caveat_ws.iter_rows(min_row=2):
        row[1].alignment = wrap

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def write_all(report: BacktestReport, out_dir: Path) -> list[Path]:
    """Everything: xlsx (if possible), CSVs, markdown and the raw JSON."""
    out_dir.mkdir(parents=True, exist_ok=True)
    written = list(write_csvs(report, out_dir))

    xlsx = write_workbook(report, out_dir / "backtest.xlsx")
    if xlsx:
        written.insert(0, xlsx)

    md = out_dir / "backtest.md"
    md.write_text(render_markdown(report), encoding="utf-8")
    written.append(md)

    raw = out_dir / "backtest.json"
    raw.write_text(
        json.dumps(report.model_dump(mode="json"), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    written.append(raw)

    return written
